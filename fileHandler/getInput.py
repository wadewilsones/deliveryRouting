#This file contains handling data import from excel and adding it to HashTable
import openpyxl;
import dataStructures.Package;
import dataStructures.packagesHashTable;

# __init__.py


def insertData(newTable, type):


        #open file with data
            excelData = openpyxl.load_workbook(f'./fileHandler/{type}.xlsx')
            excelSheet = excelData['Sheet1']

            if(type == "package"):
                #Insert each row into Package object
                for row in excelSheet.iter_rows(values_only=True):

                    # load only first 8 columns with slice
                    columns = row[:8] 

                    #Create a new package with unpacking values
                    package = dataStructures.Package.Package(*columns)

                    #Add package to HashTable
                    newTable.insert(package)


            #Add data to graph vertex
            elif(type == "addresses"):
                index = 0
                for row in excelSheet.iter_rows(values_only=True):
                    column = row[:1]
                    if None in column:
                        continue  # skip row with None values
                    print(column)
                    newTable.updateVertex(index, column)
                    index += 1
                    
                

            #Add data to graph edge
            elif(type == "distance"):
                index = 27

                for row in  excelSheet.iter_rows(values_only=True):
                 column = row[:index]

                 for cell in row:
                    #Add cell into vertex if the distance not null
                    if cell is not None:
                        vertexIn = 0
                        for vertex in newTable.vertices:
                            vertexInsecond = 1
                            for vertexSecond in newTable.vertices:
                                newTable.updateEdge(newTable.vertices[vertexIn],newTable.vertices[vertexInsecond], int(cell)) #add edge
                                vertexInsecond += 1
                            vertexIn += 1
                print(newTable.get_distanceOfVertexes(0,1))       
                    
                index = index-1

  



   


